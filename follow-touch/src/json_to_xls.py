import os
import json
from openpyxl import Workbook
import re

class FollowTouchConversion():
    def __init__(self, json_file=None, xls_file=None):
        self.json_file = json_file
        self.xls_file = xls_file or "out.xls"
        self.parts = self._get_filename_parts()
        self.valid_filenumbers = ['0', '1', '2', '3']
 
        self.workbook = Workbook()


    def file_exist(self, path):
        return os.path.isfile(path)


    def _get_filename_parts(self):
        parts = {}
        parts['dir'] = ""
        parts['name'] = ""
        parts['number'] = ""
        parts['time'] = ""
        try:
            dirname, basename = os.path.split(self.json_file)
            print (basename)
            m = re.search("(.*)_(.*)_(.*_.*).json", basename)
            if m:
                parts['dir'] = dirname
                parts['name'] = m.group(1)
                parts['number'] = m.group(2)
                parts['time'] = m.group(3)
                return parts
        except:
            return parts
        

    def convert(self):
        '''
        This start the process of reading a json file, create a workbook and write the xls file
        '''
        for number in range(0, len(self.valid_filenumbers)):
            path = os.sep.join((self.parts.get("dir"), "%s_%s_%s.json" % (self.parts.get("name"), number, self.parts.get("time"))))
            if self.file_exist(path):
                print(path)
                self.workbook.create_sheet("%s_%s" % (self.parts.get("name"), number), int(number))

                with open(path, '+r') as fp:
                    self._add_data_to_worksheet(json.load(fp))
            
        self._write(self.workbook)
        
    def _read(self):
        if self.json_file:
            with open(self.json_file, '+r') as fp:
                self._create_workbook(json.load(fp))

    def _add_data_to_worksheet(self, json_data):
        name = list(json_data.keys())[0]
        ws = self.workbook[name]

        measurements = json_data.get(name)
        
        ws.cell(row=1, column=1, value="timestamp")
        ws.cell(row=1, column=2, value="infrared")
        ws.cell(row=1, column=3, value="sensor_0")
        ws.cell(row=1, column=4, value="sensor_1")
        ws.cell(row=1, column=5, value="sensor_2")
        ws.cell(row=1, column=6, value="sensor_3")

        row = 2
        for measurement in measurements:
            timestamp = measurement.get("t")
            timestamp_value = list(timestamp.keys())[0]
            
            ws.cell(row=row, column=1, value=int(list(timestamp.keys())[0]))

            ws.cell(row=row, column=2, value='on')
            ws.cell(row=row+1, column=2, value='off')

            sensors = timestamp.get(timestamp_value)
            for c in range(0, len(sensors)):
                column = int(sensors[c].get('s')) + 3
                ws.cell(row=row, column=column, value=int(sensors[c].get('i')))
                ws.cell(row=row+1, column=column, value=int(sensors[c].get('b')))

            row += 2

    def _write(self, wb):
        outfile = os.sep.join((self.parts.get("dir"), "%s_%s.xls" % (self.parts.get("name"), self.parts.get("time"))))
        wb.save(outfile)        


if __name__ == '__main__':
    filename = os.sep.join(('/home', 'andre', 'output', 'follow_touch_0_20240813_154021.json'))
    print(filename)
    follow_touch = FollowTouch(json_file=filename)
    follow_touch.convert()
